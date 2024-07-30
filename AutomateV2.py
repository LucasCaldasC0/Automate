import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
import os
import ctypes
import pandas_market_calendars as mcal
import tkinter as tk
from tkinter import filedialog
from tkinter import *
import streamlit as st

def selecionar_arquivo():
  
  arquivo_selecionado = filedialog.askopenfilename(title="Selecionar Arquivo CSV", filetypes=[("Arquivos CSV", "*.csv")])
  if arquivo_selecionado:
    processar_arquivo(arquivo_selecionado)
    dashboard()

def processar_arquivo(arquivo_selecionado):
   # Leitura arquivo csv e inserção de novas colunas
    
    table = pd.read_csv(arquivo_selecionado)
    novas_colunas = {
    'DataAtual': '',
    'SLA envio de SL Recrutador': '',
    'SLA envio de SL Comercial': '',
    'Dias da vaga em aberto PTC': '',
    'SLA Macro': '',
    'SLA Retorno do cliente': '',
    'Dias sem retorno do cliente': '',
    'Dias da última SL enviada recrutamento': '',
    'Dias da última SL enviada comercial': '',
    'Dias para resposta do cliente': ''}
    table = table.assign(**novas_colunas)

    # Fixação e ordenação das colunas especificas
    order_nova = ['Business Unit', 'Position Name', 'Job Location', 'Job Client','Job Stage', 'Headcount', 'Job Owner','Type of Vacancy','Nome do Recrutador','Open Date','SLA Macro','Data de alinhamento','Dias da vaga em aberto PTC','1º Data SL recrutamento','Última Data SL recrutamento','Dias da última SL enviada recrutamento','SLA envio de SL Recrutador','1º Data SL comercial','Última Data SL comercial','Dias da última SL enviada comercial','SLA envio de SL Comercial','1º Data do retorno do cliente','Dias para resposta do cliente','Última data do retorno do cliente','Dias sem retorno do cliente','SLA Retorno do cliente','Status atual do processo','Job Stage','Job Status','Observações','Minimum Salary','Maximum Salary','Job Team','DataAtual']
    colunas_totais = table.columns.tolist()
    order_total = []
    #Verificação se existem colunas fora dessa lista de ordenação e as inserem nas colunas finais caso seja true
    for col in order_nova:
        if col in colunas_totais:
            order_total.append(col)
            colunas_totais.remove(col)
    
    for col in colunas_totais:
        order_total.append(col)
    
    
    table = table[order_total]
    table.to_csv(arquivo_selecionado, index=False)
    
    
    dia_atual = pd.to_datetime('today').strftime('%Y-%m-%d')

    calendario = mcal.get_calendar('BMF')
    #########Ler o arquivo csv e converter em formato xlsx
    #download_ = os.path.join(os.path.expanduser('~'),'Downloads')
    #arquivo_ = os.path.join(download_,processar_arquivo)#all-jobs-export.csv
        
    #table = pd.read_csv(arquivo_)
    
    
    table = pd.read_csv(arquivo_selecionado)
    table.to_excel("all-jobs.xlsx", sheet_name="all-jobs", index=False)

    ############Criação da nova coluna para obtenção dos dados como subtração
    table = pd.read_excel("all-jobs.xlsx",sheet_name='all-jobs')

    ##############Transformar colunas em datetime
    table['Open Date'] = pd.to_datetime(table['Open Date'])
    table['Open Date'] = table['Open Date'].dt.date

    table['Data de alinhamento'] = pd.to_datetime(table['Data de alinhamento'])
    table['Data de alinhamento'] = table['Data de alinhamento'].dt.date

    table['1º Data SL comercial'] = pd.to_datetime(table['1º Data SL comercial'])
    table['1º Data SL comercial'] = table['1º Data SL comercial'].dt.date

    table['1º Data do retorno do cliente'] = pd.to_datetime(table['1º Data do retorno do cliente'])
    table['1º Data do retorno do cliente'] = table['1º Data do retorno do cliente'].dt.date

    table['1º Data SL recrutamento'] = pd.to_datetime(table['1º Data SL recrutamento'])
    table['1º Data SL recrutamento'] = table['1º Data SL recrutamento'].dt.date

    table['Última Data SL recrutamento'] = pd.to_datetime(table['Última Data SL recrutamento'])
    table['Última Data SL recrutamento'] = table['Última Data SL recrutamento'].dt.date

    table['Última data do retorno do cliente'] = pd.to_datetime(table['Última data do retorno do cliente'])
    table['Última data do retorno do cliente'] = table['Última data do retorno do cliente'].dt.date

    table['Última Data SL comercial'] = pd.to_datetime(table['Última Data SL comercial'])
    table['Última Data SL comercial'] = table['Última Data SL comercial'].dt.date

    table['DataAtual'] = dia_atual
    table['DataAtual'] = pd.to_datetime(table['DataAtual'])
    table['DataAtual'] = table['DataAtual'].dt.date


    ###############Fórmula para as colunas
    ######Fórmula dias úteis SLA Macro


    for index_2, row in table.iterrows():
        open_date_1 = row['1º Data SL recrutamento']
        atual_date_1 = row['1º Data SL comercial']
        if pd.isna(atual_date_1) or pd.isna(open_date_1):
                continue
        if open_date_1 > atual_date_1:
            continue
        dias_uteis_1 = len(calendario.schedule(start_date=open_date_1,end_date=atual_date_1))
        table.loc[index_2,'SLA envio de SL Comercial'] = dias_uteis_1

    for index, row in table.iterrows():
        open_date = row['Open Date']
        atual_date = row['DataAtual']
        if pd.isna(open_date) or pd.isna(atual_date):
            continue
        if open_date > atual_date:
            continue
        dias_uteis = len(calendario.schedule(start_date=open_date,end_date=atual_date))
        table.loc[index,'SLA Macro'] = dias_uteis

    ######Fórmula dias úteis SLA Recrutador
    for index_1, row in table.iterrows():
        open_date_0 = row['Data de alinhamento']
        atual_date_0 = row['1º Data SL recrutamento']
        if pd.isna(open_date_0) or pd.isna(atual_date_0):
            continue
        if open_date_0 > atual_date_0:
            continue    
        dias_uteis_0 = len(calendario.schedule(start_date=open_date_0,end_date=atual_date_0))
        table.loc[index_1,'SLA envio de SL Recrutador'] = dias_uteis_0

        ######Fórmula dias úteis SLA Cliente

    ######Fórmula dias úteis SLA Retorno do cliente
    for index_3, row in table.iterrows():
        open_date_2 = row['1º Data SL comercial']
        atual_date_2 = row['1º Data do retorno do cliente']
        if pd.isna(open_date_2) or pd.isna(atual_date_2):
            continue
        if open_date_2 > atual_date_2:
            continue
        dias_uteis_2 = len(calendario.schedule(start_date=open_date_2,end_date=atual_date_2))
        table.loc[index_3,'SLA Retorno do cliente'] = dias_uteis_2

    ######Fórmula dias abertura da vaga
    for index_4, row in table.iterrows():
        open_date_3 = row['Data de alinhamento']
        atual_date_3 = row['DataAtual']
        if pd.isna(open_date_3) or pd.isna(atual_date_3):
            continue
        if open_date_3 > atual_date_3:
            continue
        dias_uteis_3 = len(calendario.schedule(start_date=open_date_3,end_date=atual_date_3))
        table.loc[index_4,'Dias da vaga em aberto PTC'] = dias_uteis_3
    
    for index_5, row in table.iterrows():
        open_date_4 = row['Última data do retorno do cliente']
        atual_date_4 = row['DataAtual']
        if pd.isna(open_date_4) or pd.isna(atual_date_4):
            continue
        if open_date_4 > atual_date_4:
            continue
        dias_uteis_4 = len(calendario.schedule(start_date=open_date_4,end_date=atual_date_4))
        table.loc[index_5,'Dias sem retorno do cliente'] = dias_uteis_4

    for index_6, row in table.iterrows():
        open_date_5 = row['Última Data SL recrutamento']
        atual_date_5 = row['DataAtual']
        if pd.isna(open_date_5) or pd.isna(atual_date_5):
            continue
        if open_date_5 > atual_date_5:
            continue
        dias_uteis_5 = len(calendario.schedule(start_date=open_date_5,end_date=atual_date_5))
        table.loc[index_6,'Dias da última SL enviada recrutamento'] = dias_uteis_5

    for index_7, row in table.iterrows():
        open_date_6 = row['Última Data SL comercial']
        atual_date_6 = row['DataAtual']
        if pd.isna(open_date_6) or pd.isna(atual_date_6):
            continue
        if open_date_6 > atual_date_6:
            continue
        dias_uteis_6 = len(calendario.schedule(start_date=open_date_6,end_date=atual_date_5))
        table.loc[index_7,'Dias da última SL enviada comercial'] = dias_uteis_6

    for index_8, row in table.iterrows():
        open_date_7 = row['Última Data SL comercial']
        atual_date_7 = row['Última data do retorno do cliente']
        if pd.isna(open_date_7) or pd.isna(atual_date_7):
            continue
        if open_date_7 > atual_date_7:
            continue
        dias_uteis_7 = len(calendario.schedule(start_date=open_date_7,end_date=atual_date_7))
        table.loc[index_8,'Dias para resposta do cliente'] = dias_uteis_7

    table.to_excel("all-jobs.xlsx", sheet_name="all-jobs", index=False)

    #############transformar em tabela e adicionar cor nos valores especificados
    time.sleep(2)
    wb = load_workbook('all-jobs.xlsx')
    ws = wb.active

    def get_last_column_ref(max_column_index):
        """Converts a maximum column index (e.g., 27) to a multi-letter Excel column reference (e.g., AA)."""
        col_ref = ""
        while max_column_index >= 0:
            remainder = max_column_index % 26
            dividend = max_column_index // 26
            col_ref = chr(ord('A') + remainder) + col_ref
            max_column_index = dividend - 1

        return col_ref

    last_col_ref = get_last_column_ref(ws.max_column - 1)

    first_row = ws.min_row
    last_row = ws.max_row

    table_ref = f'A{first_row}:{last_col_ref}{last_row}'
    table_mod = Table(displayName='Table_2',ref=f'{table_ref}')
    style = TableStyleInfo(name='TableStyleMedium6',showColumnStripes=True,showFirstColumn=False,showLastColumn=False,showRowStripes=True)
    table_mod.tableStyleInfo = style
    ws.add_table(table_mod)

    col_index = 0
    col_index_1 = 0
    col_index_2 = 0
    col_index_3 = 0
    col_index_4 = 0


    for cell in ws[1]:
        if cell.internal_value == 'SLA Macro':
            col_index = cell.col_idx
            break
    for cell in ws[1]:
        if cell.internal_value == 'SLA envio de SL Recrutador':
            col_index_1 = cell.col_idx
            break
    for cell in ws[1]:
        if cell.internal_value == 'SLA envio de SL Comercial':
            col_index_2 = cell.col_idx
            break
    for cell in ws[1]:
        if cell.internal_value == 'SLA Retorno do cliente':
            col_index_3 = cell.col_idx
            break
    for cell in ws[1]:
        if cell.internal_value == 'Dias da vaga em aberto PTC':
            col_index_4 = cell.col_idx
            break
    for cell in ws[1]:
        if cell.internal_value == 'Dias sem retorno do cliente':
            col_index_5 = cell.col_idx
            break
    for cell in ws[1]:
        if cell.internal_value == 'Dias da última SL enviada recrutamento':
            col_index_6 = cell.col_idx
            break
    for cell in ws[1]:
        if cell.internal_value == 'Dias da última SL enviada comercial':
            col_index_7 = cell.col_idx
            break
    for cell in ws[1]:
        if cell.internal_value == 'Dias para resposta do cliente':
            col_index_8 = cell.col_idx
            break

    green = '32FF54'
    yellow = 'FAFF52'
    red = 'FF2424'

    for cells_in_row in ws.iter_rows(min_row=2,min_col=col_index, max_col=col_index):
            if cells_in_row[0].internal_value is None or cells_in_row[0].internal_value == '':
                continue
            if cells_in_row[0].internal_value < 8:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=green)
            elif cells_in_row[0].internal_value > 7 and cells_in_row[0].internal_value <= 15:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=yellow)
            elif cells_in_row[0].internal_value > 15:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=red)

    for cells_in_row in ws.iter_rows(min_row=2,min_col=col_index_1, max_col=col_index_1):
            if cells_in_row[0].internal_value is None or cells_in_row[0].internal_value == '':
                continue
            if cells_in_row[0].internal_value < 8:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=green)
            elif cells_in_row[0].internal_value > 7 and cells_in_row[0].internal_value <= 15:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=yellow)
            elif cells_in_row[0].internal_value > 15:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=red)
            
    for cells_in_row in ws.iter_rows(min_row=2,min_col=col_index_2, max_col=col_index_2):
            if cells_in_row[0].internal_value is None or cells_in_row[0].internal_value == '':
                continue
            if cells_in_row[0].internal_value < 5:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=green)
            elif cells_in_row[0].internal_value > 4 and cells_in_row[0].internal_value <= 7:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=yellow)
            elif cells_in_row[0].internal_value > 7:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=red)

    for cells_in_row in ws.iter_rows(min_row=2,min_col=col_index_3, max_col=col_index_3):
            if cells_in_row[0].internal_value is None or cells_in_row[0].internal_value == '':
                continue
            if cells_in_row[0].internal_value < 8:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=green)
            elif cells_in_row[0].internal_value > 7 and cells_in_row[0].internal_value <= 15:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=yellow)
            elif cells_in_row[0].internal_value > 15:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=red)

    for cells_in_row in ws.iter_rows(min_row=2,min_col=col_index_4, max_col=col_index_4):
            if cells_in_row[0].internal_value is None or cells_in_row[0].internal_value == '':
                continue
            if cells_in_row[0].internal_value < 8:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=green)
            elif cells_in_row[0].internal_value > 7 and cells_in_row[0].internal_value <= 15:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=yellow)
            else:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=red)

    for cells_in_row in ws.iter_rows(min_row=2,min_col=col_index_5, max_col=col_index_5):
            if cells_in_row[0].internal_value is None or cells_in_row[0].internal_value == '':
                continue
            if cells_in_row[0].internal_value < 8:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=green)
            elif cells_in_row[0].internal_value > 7 and cells_in_row[0].internal_value <= 15:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=yellow)
            else:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=red)

    for cells_in_row in ws.iter_rows(min_row=2,min_col=col_index_6, max_col=col_index_6):
            if cells_in_row[0].internal_value is None or cells_in_row[0].internal_value == '':
                continue
            if cells_in_row[0].internal_value < 8:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=green)
            elif cells_in_row[0].internal_value > 7 and cells_in_row[0].internal_value <= 15:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=yellow)
            else:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=red)

    for cells_in_row in ws.iter_rows(min_row=2,min_col=col_index_7, max_col=col_index_7):
            if cells_in_row[0].internal_value is None or cells_in_row[0].internal_value == '':
                continue
            if cells_in_row[0].internal_value < 5:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=green)
            elif cells_in_row[0].internal_value > 4 and cells_in_row[0].internal_value <= 7:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=yellow)
            else:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=red)

    for cells_in_row in ws.iter_rows(min_row=2,min_col=col_index_8, max_col=col_index_8):
            if cells_in_row[0].internal_value is None or cells_in_row[0].internal_value == '':
                continue
            if cells_in_row[0].internal_value < 8:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=green)
            elif cells_in_row[0].internal_value > 7 and cells_in_row[0].internal_value <= 15:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=yellow)
            else:
                cells_in_row[0].fill = PatternFill(patternType='solid', fgColor=red)


    wb.save('all-jobs.xlsx')
    wb.close()

    os.remove(arquivo_selecionado)

    ctypes.windll.user32.MessageBoxW(0, "O Script foi executado com sucesso!!", "Automação", 64)

def dashboard():
    st.title("Teste de funcionamento")



janela = tk.Tk()
janela.title("Automação")
janela.geometry('300x120')

texto_orientar = Label(janela,text=f'''Clique no botão para selecionar o arquivo em 
formato .CSV''')
texto_orientar.grid(column=0,row=0, padx=10, pady=10)

botao_selecionar_arquivo = tk.Button(janela,text="Selecionar arquivo .csv",command=selecionar_arquivo)
botao_selecionar_arquivo.grid(column=0,row=2, padx=80, pady=10)
#botao_selecionar_arquivo.pack(pady=40)

janela.mainloop()